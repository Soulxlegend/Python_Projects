import openpyxl
# from hashing import encrypting, decrypting 
import pwinput

class ExcelReader:
    def __init__(self):
        # self.file_path = "Book1.xlsx"
        self.workbook = openpyxl.load_workbook('D:\Dev\Web\Python_Projects\Book1.xlsx')
        self.sheet = self.workbook['Sheet1']
    def create_account(self):
        # Create a new account
        for i in  range(1, 500):
            if  self.sheet.cell(row=i+1, column=1).value == None:
                try:
                    rules = """
                        This is Where you have to enter your personal details so We can create your bank account.
                        Please provide the following information:
                        1. Name
                        2. Email
                        3. Phone Number
                        4. Address
                        5. Account Type (Savings or Current)
                        6. Initial Deposit
                        11. Account Pin
                        12. Account Password
                        13. Account Security Question
                        14. Account Security Answer
                        """
                    print(rules)
                    
                    Account_holder_name = input("Enter Your Full Name: ")
                    Father_name = input("Please Enter your Father's Full Name: ")
                    Email = input("Enter your Email address: ")
                    phone_no = int(input("Enter your Phone Number: "))
                    Permanent_address = input("Enter your Permanent Address: ")
                    Initial_deposit = int(input("Your Initial Deposit in Int: "))
                    Account_type = input("Enter your Account Type (Savings or Current): ")
                    Account_password = pwinput.pwinput(prompt='Enter your password: ', mask='•')
                    Account_pin = pwinput.pwinput(prompt='Enter your PIN: ', mask='*')
                    Account_security_question = input("Enter your Account Security Question: ")
                    Account_security_answer = input("Enter your Account Security Answer: ")

                    if Account_holder_name == "" or Father_name == "" or Email == "" or phone_no == "" or Permanent_address == "" or Initial_deposit == 0 or Account_type == "" or Account_pin == 0 or Account_password == "" or Account_security_question == "" or Account_security_answer == "": 
                        print("Please fill all the fields")

                    elif len(Account_pin) < 6:
                        print("Please enter a 6 digit PIN")
                        
                    elif len(Account_password) < 8:
                        print("Your Password must contain at least 8 characters")

                    elif Account_password != Account_password.isalnum():
                        print("Your Password must contain at least one alphabet and one number")
                        
                    else:
                        self.sheet.cell(row=i+1,column=1).value = Account_holder_name
                        self.sheet.cell(row=i+1,column=2).value = Father_name
                        self.sheet.cell(row=i+1,column=3).value = Email
                        self.sheet.cell(row=i+1,column=4).value = phone_no
                        self.sheet.cell(row=i+1,column=5).value = Permanent_address
                        self.sheet.cell(row=i+1,column=6).value = Initial_deposit
                        self.sheet.cell(row=i+1,column=7).value = Account_type
                        self.sheet.cell(row=i+1,column=8).value = Account_pin
                        self.sheet.cell(row=i+1,column=9).value = Account_password
                        self.sheet.cell(row=i+1,column=10).value = Account_security_question
                        self.sheet.cell(row=i+1,column=11).value = Account_security_answer
                        def account_numgen():
                            import random
                            x = '65764323'
                            for i in range(8):
                                random_numbers = str(random.randint(0, 9))
                                x += random_numbers
                            return x
                        self.sheet.cell(row=i+1,column=12).value = account_numgen()
                        print(f"Our module have assigned this account number {self.sheet.cell(row=i+1,column=12).value} to you! please save this Account Number for further login or transaction! ")
                        self.workbook.save("D:\Dev\Web\Python_Projects\Book1.xlsx")
                        print("Account Created Successfully!")
                except Exception as e:
                    print(f"Invalid Argument! {e}")

    def UI():
        starting_ui = """
        =====================================================
            Please select one to proceed:
            1. Create an Account.
            2. Open Account INFO.
            3. Deposit money into the Account.
            4. Withdraw money from the Account.
            5. Exit.
        =====================================================
        """
        print(starting_ui)

        try:
            choice = int(input("Enter your choice: "))
        except:
            print("Invalid choice. Please enter a number between 1 and 5.")

        if choice == 1:
           ExcelReader().create_account()
        else:
            print("Invalid choice. Please enter a number between 1 and 5.")

        # elif choice == 2:

        # elif choice == 3:
        
        # elif choice == 4:

        # elif choice == 5:
            # print("Exiting...")
            # exit()


    # def check_name():





    # def accounts():
        

    

    # def send_pass():



    # def get_pass():


if __name__ == "__main__":
    ExcelReader.UI()


# import pwinput

# password = pwinput.pwinput(prompt='Enter your password: ', mask='•')
# print("Your password is:", password)



#the imitation game
